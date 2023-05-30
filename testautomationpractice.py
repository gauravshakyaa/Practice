import datetime
import os
import time

from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from datetime import date
import openpyxl
from Practice import find
import mysql.connector


# ops.add_argument("--disable-notifications")  # to disable alert of camera and microphone

def init():
    serv_obj = Service("D:\Drivers\chromedriver_win32\chromedriver.exe")

    # preferences = {"download.default_directory": "D:\QA Materials"}
    ops = webdriver.ChromeOptions()
    # ops.headless = True
    # ops.add_experimental_option("prefs", preferences)
    # ops.add_argument('--headless=new')
    driver = webdriver.Chrome(service=serv_obj, options=ops)
    return driver


driver = init()
# opt = Options()
# opt.add_experimental_option("debuggerAddress", "localhost:8989")
# driver = webdriver.Chrome(executable_path="D:\Drivers\chromedriver_win32",chrome_options=opt)

# driver.get('https://testautomationpractice.blogspot.com/')
# driver.maximize_window()

actions = ActionChains(driver)


def search_bar_send_selenium_open_links():
    search = driver.find_element(By.CLASS_NAME, "wikipedia-search-input")

    search.send_keys('selenium')

    search.submit()
    time.sleep(1)
    links = driver.find_elements(By.XPATH, "//div[@id='wikipedia-search-result-link']//a")

    for link in links:
        link.click()
        time.sleep(2)

    windows = driver.window_handles
    parent_window = windows[0]

    driver.switch_to.window(parent_window)


def speed_dropdown():
    driver.get("https://testautomationpractice.blogspot.com/")
    driver.maximize_window()
    drp = Select(driver.find_element(By.ID, "speed"))
    drp.select_by_value('Slow')


def file_dropdown():
    drp = Select(driver.find_element(By.ID, 'files'))
    drp.select_by_value('3')


def dbl_click():
    dblclick = driver.find_element(By.CSS_SELECTOR, "button[ondblclick='myFunction1()']")
    actions.double_click(dblclick).perform()


def drag_drop():
    source = driver.find_element(By.ID, "draggable")
    target = driver.find_element(By.ID, "droppable")
    actions.drag_and_drop(source, target).perform()

    print(target.text)


def slider():
    slider = driver.find_element(By.ID, 'slider')

    actions.move_to_element(slider).click_and_hold().move_by_offset(50, 50).release().perform()
    print(slider.location)  # {'x': 1179, 'y': 742}
    print(slider.location_once_scrolled_into_view)  # {'x': 1179, 'y': 1}
    actions.context_click(slider)


def alert():
    driver.find_element(By.CSS_SELECTOR, "button[onclick='myFunction()']").click()
    alert = driver.switch_to.alert
    alert.dismiss()


def frame():
    driver.get("https://www.selenium.dev/selenium/docs/api/java/index.html?overview-summary.html")
    driver.maximize_window()
    time.sleep(3)

    driver.switch_to.frame('packageListFrame')
    driver.find_element(By.LINK_TEXT, "org.openqa.selenium.bidi").click()

    time.sleep(4)

    # driver.switch_to.parent_frame()
    driver.switch_to.default_content()
    driver.switch_to.frame('packageFrame')

    driver.find_element(By.XPATH, "//a[@title='interface in org.openqa.selenium.bidi']").click()


def scrollParty():
    driver.get("https://web.karobarapp.com/login")
    driver.maximize_window()

    driver.find_element(By.XPATH, '//input[@type="text" and @placeholder="Enter your Phone Number" and '
                                  '@class="input"]').send_keys('9860725577')
    time.sleep(2)
    # Find the submit button locator and click it
    # driver.find_element(By.CLASS_NAME, '.button').click()
    driver.find_element(By.XPATH, '//*[@id="app"]/div/section/div[2]/div/div[1]/div/button').click()
    time.sleep(2)
    driver.find_element(By.CLASS_NAME, 'otp-input').send_keys(
        '123456')  # Send otp code, automatically pass through login page
    time.sleep(2)
    # Select admin account and click it
    driver.find_element(By.XPATH, "//small[normalize-space()='Admin']").click()
    time.sleep(4)

    driver.get("https://web.karobarapp.com/parties")
    time.sleep(2)
    party = driver.find_element(By.XPATH, '//*[@id="app"]/div[2]/section[2]/div/div[1]/article/div[3]')

    # driver.execute_script("document.querySelector('.content').scrollTop=500") #  scroll with height

    # driver.execute_script("window.scrollBy(0, document.body.scrollHeight)") #  scroll to bottom

    # driver.execute_script("arguments[0].scrollIntoView();", flag) #  scroll unless you see specific element
    driver.execute_script("document.querySelector('article.tile.is-child.notification.is-white .content').scrollHeight")

    time.sleep(2)


def twitterscroll():
    driver.get("https://twitter.com/")
    driver.maximize_window()
    time.sleep(3)
    previous = driver.execute_script("return document.body.scrollHeight")
    print(previous)
    time.sleep(3)
    while True:
        driver.execute_script("window.scroll(0, document.body.scrollHeight)")
        time.sleep(3)
        new = driver.execute_script("return document.body.scrollHeight")
        if new == previous:
            break
        previous = new


def scrollbykeys():
    driver.get("https://twitter.com/")
    driver.maximize_window()
    time.sleep(3)
    element = driver.find_element(By.TAG_NAME, "body")
    while True:
        element.send_keys(Keys.PAGE_DOWN)
        time.sleep(3)


def scrollbyelement():
    driver.get("https://qavbox.github.io/demo/webtable/")
    driver.maximize_window()
    time.sleep(3)

    driver.execute_script("return document.getElementById('table02')")

    # # driver.execute_script("argument[0].scrollIntoView);", element)
    # time.sleep(3)
    #
    # while True:
    #     driver.execute_script("document.getElementById('table02').scrollTop=500")
    #
    #     new = driver.execute_script("return document.getElementById('table02')")
    #     if previous == new:
    #         break
    #
    #     previous = new


def notificationpopup():
    driver.get("https://whatmylocation.com/")
    driver.maximize_window()


def webtable():
    driver.get("https://testautomationpractice.blogspot.com")

    driver.find_element(By.NAME, "BookTable")

    rows = len(driver.find_elements(By.XPATH, "//table[@name='BookTable']//tr"))
    cols = len(driver.find_elements(By.XPATH, "//table[@name='BookTable']//tr//th"))

    print(rows, cols)

    # for r in range(2, rows + 1):
    #     for c in range(1, cols + 1):
    #         data = driver.find_element(By.XPATH, "//table[@name='BookTable']//tbody//tr["+str(r)+"]//td["+str(c)+"]").text
    #         print(data)

    for r in range(2, rows + 1):
        for c in range(1, cols + 1):
            user = driver.find_element(By.XPATH, "//table[@name='BookTable']//tbody//tr[" + str(r) + "]//td[2]").text
            # print(user)
            if user == "Mukesh":
                book = driver.find_element(By.XPATH,
                                           "//table[@name='BookTable']//tbody//tr[" + str(r) + "]//td[1]").text
                print(book)
                break


def dynamic_table_orangehrm():
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    driver.maximize_window()
    time.sleep(4)
    driver.find_element(By.NAME, "username").send_keys("Admin")
    driver.find_element(By.NAME, "password").send_keys("admin123")
    driver.find_element(By.CSS_SELECTOR, ".oxd-button").click()
    time.sleep(3)
    # wait = WebDriverWait(driver, timeout=5, poll_frequency=2)
    # wait.until(EC.visibility_of_element_located(driver.find_element(By.XPATH, "//span[text()='Admin']")))
    driver.find_element(By.XPATH, "//span[text()='Admin']").click()
    time.sleep(3)
    # driver.find_element(By.XPATH, "ul[class='oxd-dropdown-menu']li").click()
    # time.sleep(3)
    rows = len(driver.find_elements(By.XPATH,
                                    "//div[@class='oxd-table']//div[@class='oxd-table-body']//div[@class='oxd-table-card']"))
    cols = len(driver.find_elements(By.XPATH,
                                    "//div[@class='oxd-table']//div[@class='oxd-table-header']//div[@role='row']//div[@role='columnheader']"))
    print(type(cols))
    print("Rows= ", rows, "Columns= ", cols)

    for r in range(1, rows + 1):
        for c in range(2, cols):
            # all_data = driver.find_element(By.XPATH,
            #                                "//div[@class='oxd-table']//div[@class='oxd-table-body']//div[@class='oxd-table-card'][" + str(
            #                                    r) + "]").text
            all_data = driver.find_element(By.XPATH,
                                           "//div[@class='oxd-table']//div[@class='oxd-table-body']//div[@class='oxd-table-card'][" + str(
                                               r) + "]/child::div/child::div[5]").text

            if all_data == "Disabled":
                dis_user = driver.find_element(By.XPATH,
                                               "//div[@class='oxd-table']//div[@class='oxd-table-body']//div[@class='oxd-table-card'][" + str(
                                                   r) + "]/child::div/child::div[4]").text
                print(dis_user)
                break


def datepicker():
    driver.get("https://jqueryui.com/datepicker/")
    driver.maximize_window()
    time.sleep(4)

    driver.switch_to.frame(0)

    driver.find_element(By.XPATH, "//input[@id='datepicker']").click()

    year = 2029
    month = "December"
    day = 30
    mon = driver.find_element(By.XPATH, "(//span[@class='ui-datepicker-month'])").text
    yr = driver.find_element(By.XPATH, "(//span[@class='ui-datepicker-year'])").text
    current_year = datetime.datetime.now().year
    right_arrow = driver.find_element(By.XPATH, "//span[@class='ui-icon ui-icon-circle-triangle-e']")
    left_arrow = driver.find_element(By.XPATH, "//span[@class='ui-icon ui-icon-circle-triangle-w']")

    while True:

        if mon == month and yr == year:
            break

        elif year > int(yr):
            right_arrow.click()

        elif year < int(yr):
            left_arrow.click()

        else:
            pass

    days = driver.find_elements(By.XPATH, "//table[@class='ui-datepicker-calendar']//tbody//td")

    for i in days:
        if i.text == day:
            i.click()
            break

    time.sleep(4)


def datepicker2():
    driver.get("https://www.dummyticket.com/dummy-ticket-for-visa-application/")
    driver.maximize_window()
    time.sleep(4)
    driver.execute_script("window.scrollTo(0, document.body.scrollTop=500)")  # scroll to view the datepicker
    driver.find_element(By.XPATH, "//input[@id='dob']").click()
    date_picker = Select(driver.find_element(By.XPATH, "//select[@aria-label='Select month']"))
    date_picker.select_by_visible_text("Dec")


def dateandtime():
    x = datetime.datetime.now()
    print(x.year)
    print(x.strftime("%C"))


def keyboard_actions():
    driver.get("https://text-compare.com/")
    driver.maximize_window()

    input1 = driver.find_element(By.ID, "inputText1")
    input2 = driver.find_element(By.ID, "inputText2")

    input1.send_keys("Welcome")

    actions.key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
    actions.key_down(Keys.CONTROL).send_keys("c").key_up(Keys.CONTROL).perform()

    actions.key_down(Keys.TAB).perform()

    actions.key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()


def download_upload():
    # preferences = {"download.default_directory": "D:\QA Materials"}
    # ops = webdriver.ChromeOptions()
    # ops.add_experimental_option("prefs", preferences)
    driver.get("https://file-examples.com/index.php/sample-documents-download/")
    driver.maximize_window()

    driver.find_element(By.XPATH,
                        "//a[@href='https://file-examples.com/index.php/sample-documents-download/sample-doc-download/']").click()
    time.sleep(2)


def bootstrap_span_select():
    driver.get("https://www.dummyticket.com/dummy-ticket-for-visa-application/")
    driver.maximize_window()

    driver.find_element(By.XPATH, "//span[@id='select2-billing_country-container']").click()


def screenshot():
    driver.get("https://demo.nopcommerce.com/")
    driver.maximize_window()
    # driver.save_screenshot(os.getcwd()+"//homepage.png")
    # driver.find_element(By.XPATH, "//span[@id='select2-billing_country-container']").click()
    # driver.get_screenshot_as_file(os.getcwd()+"//home.png")
    # driver.get_screenshot_as_png()
    driver.get_screenshot_as_file("home.png")


def switch_tab():
    driver.get("https://demo.nopcommerce.com/")
    driver.switch_to.new_window()
    driver.get("https://demo.nopcommerce.com/cart")
    windows = driver.window_handles
    for window in windows:
        print(window)

    parent_window = windows[0]

    # driver.switch_to.window(parent_window)

    # regilink = Keys.CONTROL + Keys.RETURN
    # driver.find_element(By.XPATH, "//a[normalize-space()='Register']").send_keys(regilink)

    #  Selenium 4

    time.sleep(2)


def cookies():
    driver.get("https://web.karobarapp.com/")

    cookiee = driver.get_cookies()
    # for c in cookie:
    #     print(c.get("domain"), c.get("secure"))
    print(len(cookiee))
    driver.add_cookie({"name": "MyCookie", "value": "123456"})
    cookie = driver.get_cookies()
    print(len(cookie))

    driver.delete_cookie("MyCookie")
    cookie = driver.get_cookies()
    print(len(cookie))

    driver.delete_all_cookies()
    cookie = driver.get_cookies()
    print(len(cookie))


def headless():
    driver.get("https://web.karobarapp.com/")
    print(driver.title)
    print(driver.current_url)


def excel():
    file = "C://Karobar_Products_Import_Sample.xlsx"
    workbook = openpyxl.load_workbook(file)
    sheet = workbook["Sheet1"]

    row = sheet.max_row
    cols = sheet.max_column

    for r in range(1, row + 1):
        for c in range(1, cols + 1):
            print(sheet.cell(r, c).value, end='     ')
        print()


def writeInExcel():
    file = "D://data.xlsx"
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active  # (or) sheet = workbook['sheet_name'] --get active sheet, only one sheet

    #  Header
    sheet.cell(1, 1).value = 'S.N.'
    sheet.cell(1, 2).value = 'Name'
    sheet.cell(1, 3).value = 'Address'

    #  Data
    sheet.cell(2, 1).value = 1
    sheet.cell(2, 2).value = 'Gaurav'
    sheet.cell(2, 3).value = 'Boudha'

    workbook.save(file)


def xpath():
    driver.get("https://the-internet.herokuapp.com/challenging_dom")
    driver.maximize_window()
    self = find.by_element_xpath("//td[normalize-space()='Definiebas4']")


def database():
    try:
        insert_query = "INSERT INTO employee VALUES (105, 'Gaurav Shakya')"
        update_query = "UPDATE employee set full_name = 'Saurav' where id = 1;"
        delete_query = "DELETE from employee where id = 1;"

        con = mysql.connector.connect(host="localhost", port=3306, user="root", passwd="root", database="automation")
        curs = con.cursor()
        curs.execute("select *from employee")
        for row in curs:
            id_value = row[0]
            name = row[1]
            print(name)
            curs.execute(f"INSERT INTO employee VALUES ({id_value}, {name}")
            con.commit()

        print("Finished....")
        con.close()
    except:
        print("Connection unsuccessfull...")


database()

time.sleep(3)
driver.close()
driver.quit()
