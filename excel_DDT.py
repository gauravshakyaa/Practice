import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColsCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readDate(file, sheetName, row, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row, col).value


def writeData(file, sheetName, row, col, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row, col).value = data
    workbook.save(file)


def fillGreenColor(file, sheetName, row, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(row, col).fill = greenFill
    workbook.save(file)


def fillRedColor(file, sheetName, row, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    fillRed = PatternFill(start_color='d91e1e', end_color='d91e1e', fill_type='solid')
    sheet.cell(row, col).fill = fillRed
    workbook.save(file)
