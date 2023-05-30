from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import openpyxl

serv_obj = Service("D:\Drivers\chromedriver_win32\chromedriver.exe")

driver = webdriver.Chrome(service=serv_obj)


def read_all(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    row_count = sheet.max_row
    col_count = sheet.max_column
    for r in range(1, row_count + 1):
        for c in range(2, col_count + 1):
            return sheet.cell(r, c).value
    # return print(col_count)


def read_data(file, sheet, row, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    url = sheet.cell(row, col).value

    return print(url)


# rows("D://URL_Web.xlsx", "Sheet1")
# cols("D://URL_Web.xlsx", "Sheet1")
# read_data("D://URL_Web.xlsx", "Sheet1", 1, 2)

read = read_all("D://URL_Web.xlsx", "Sheet1")
